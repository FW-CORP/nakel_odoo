#!/usr/bin/env python3
"""
Flujo: export viejo (CSV/XLSX) + **extracción MSSQL** → SQLite legible.

1. **Vendedor:** el ID del CSV se cruza con `VENDEDORES` → se guarda el **nombre**
   (`NOMBRE_VENDEDOR`) y código de vendedor, no solo el número.
2. **Cliente:** el ID se cruza con `CLIENTES` → **razón social**, nombre de fantasía,
   CUIT, dirección, email, código de cliente (en Gestion no hay columnas separadas
   nombre/apellido; el dato “humano” principal es `RAZON_SOCIAL`).
3. **Artículo:** el código entero del export se convierte al **mismo formato que Odoo**
   (`default_code` con decimales, p.ej. `885725` → `8857.25`) en `codigo_articulo_odoo`;
   desde MSSQL se añaden `cod_articulo_mssql` y `descripcion_articulo`, y datos de ficha
   (`mssql_unidad_medida`, `mssql_unid_bulto`, …) más `cantidad_pedida_contexto` para
   interpretar la columna «Cantidad pedida» (ver `MAPEO_PREVENTAS_MSSQL_MASTER18.md`).
4. **Correcciones (JSON):** si el export “come” mal el entero (p. ej. `124309`→`1243.09`
   en vez de Kinder Bueno `1243.90`), `correcciones_codigo_articulo_preventas.json` corrige
   hacia el `default_code` real. Se guardan `codigo_odoo_antes_correccion` y
   `correccion_codigo_detalle`. Si la descripción MSSQL empieza por ZZ/ZZZ, `alerta_articulo_mssql`.

Opcional: `--resolver-odoo` rellena ids de Odoo (ver `pedido_sqlite_odoo.py`).

Columnas esperadas en el CSV (equivalente A–F):
  - Operacion, Vendedor, Cliente, Ruta, Codigo articulo, Cantidad pedida, fecha/hora…

Uso:
  python3 enriquecer_pedido_csv_sqlite_mssql.py \\
    --csv "/ruta/Pedidos.csv" \\
    --db /ruta/salida/pedido_omar.sqlite

Requiere: pyodbc, ODBC Driver 18, contenedor MSSQL accesible (config_nakel.MSSQL_CONFIG).

Opcional XLSX: pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import sqlite3
import sys
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, "/media/klap/raid5/cursor_files")
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from correccion_codigos_preventas import (  # noqa: E402
    DEFAULT_CORRECCIONES_JSON,
    aplicar_correcciones_codigo,
    alerta_por_descripcion_mssql,
    cargar_correcciones,
)

try:
    import pyodbc  # type: ignore
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Instala pyodbc y el ODBC Driver 18 para SQL Server."
    ) from exc

try:
    from config_nakel import MSSQL_CONFIG  # type: ignore
except ImportError as e:
    raise SystemExit(
        "Falta config_nakel.py en /media/klap/raid5/cursor_files"
    ) from e


def articulo_csv_a_default_code(raw: Any) -> str:
    """Misma lógica que inyectar_pedidos_csv_master18."""
    if raw is None or raw == "":
        return ""
    s = str(raw).strip().replace(",", ".")
    if not s:
        return ""
    if "." in s:
        try:
            d = Decimal(s)
            return format(d.normalize(), "f")
        except InvalidOperation:
            return s
    try:
        n = int(Decimal(s))
    except (InvalidOperation, ValueError):
        return s
    entero = n // 100
    frac = n % 100
    return f"{entero}.{frac:02d}"


def normalize_code(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace(",", ".")


def conectar_mssql():
    c = MSSQL_CONFIG
    conn_s = (
        f"DRIVER={{{c['driver']}}};SERVER={c['server']};DATABASE={c['database']};"
        f"UID={c['username']};PWD={c['password']};TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn_s, timeout=c.get("timeout", 30))


def detect_vendedor_columns(cur) -> tuple[str, str | None]:
    """(columna_nombre, columna_codigo_opcional) para tabla VENDEDORES."""
    cur.execute(
        """
        SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'VENDEDORES'
          AND COLUMN_NAME IN ('NOMBRE_VENDEDOR', 'NOMBRE')
        """
    )
    names = {row[0] for row in cur.fetchall()}
    nom = "NOMBRE_VENDEDOR" if "NOMBRE_VENDEDOR" in names else (
        "NOMBRE" if "NOMBRE" in names else "NOMBRE_VENDEDOR"
    )
    cur.execute(
        """
        SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'VENDEDORES'
          AND COLUMN_NAME IN ('CODIGO_VENDEDOR', 'COD_VENDEDOR')
        """
    )
    cods = [row[0] for row in cur.fetchall()]
    cod_col = cods[0] if cods else None
    return nom, cod_col


def armar_cliente_display(
    razon: str | None, nombre_fantasia: str | None
) -> str | None:
    """Una sola línea legible para listados (razón social + fantasía si aporta)."""
    r = (razon or "").strip()
    f = (nombre_fantasia or "").strip()
    if r and f and f.upper() != r.upper():
        return f"{r} — {f}"
    return r or f or None


def _sql_str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def detect_cliente_cod_column(cur) -> str | None:
    cur.execute(
        """
        SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'CLIENTES'
          AND COLUMN_NAME IN ('COD_CLIENTE', 'CODIGO_CLIENTE')
        """
    )
    names = {row[0] for row in cur.fetchall()}
    if "COD_CLIENTE" in names:
        return "COD_CLIENTE"
    if "CODIGO_CLIENTE" in names:
        return "CODIGO_CLIENTE"
    return None


def cargar_mapa_vendedores(
    cur, ids: set[int], col_nombre: str, col_codigo: str | None
) -> dict[int, tuple[str | None, str | None]]:
    if not ids:
        return {}
    qmarks = ",".join("?" * len(ids))
    cod_sql = (
        f"LTRIM(RTRIM(ISNULL(CAST({col_codigo} AS VARCHAR(20)), '')))"
        if col_codigo
        else "CAST(NULL AS VARCHAR(1))"
    )
    cur.execute(
        f"""
        SELECT ID_VENDEDOR,
               LTRIM(RTRIM(ISNULL({col_nombre}, ''))),
               {cod_sql}
        FROM VENDEDORES
        WHERE ID_VENDEDOR IN ({qmarks})
        """,
        list(ids),
    )
    out: dict[int, tuple[str | None, str | None]] = {}
    for row in cur.fetchall():
        out[int(row[0])] = (row[1] or None, row[2] or None)
    return out


def cargar_mapa_clientes(
    cur, ids: set[int], cod_col: str | None
) -> dict[int, tuple[str | None, str | None, str | None, str | None, str | None, str | None]]:
    """
    Por ID_CLIENTE: razón social, nombre fantasía, CUIT, dirección, email, código cliente.
    """
    if not ids:
        return {}
    cod_sql = (
        f"LTRIM(RTRIM(ISNULL(CAST({cod_col} AS VARCHAR(50)), '')))"
        if cod_col
        else "CAST(NULL AS VARCHAR(1))"
    )
    qmarks = ",".join("?" * len(ids))
    cur.execute(
        f"""
        SELECT ID_CLIENTE,
               LTRIM(RTRIM(ISNULL(RAZON_SOCIAL, ''))),
               LTRIM(RTRIM(ISNULL(NOMBRE_FANTASIA, ''))),
               LTRIM(RTRIM(ISNULL(CUIT, ''))),
               LTRIM(RTRIM(ISNULL(DIRECCION, ''))),
               LTRIM(RTRIM(ISNULL(EMAIL, ''))),
               {cod_sql}
        FROM CLIENTES
        WHERE ID_CLIENTE IN ({qmarks})
        """,
        list(ids),
    )
    out: dict[
        int,
        tuple[str | None, str | None, str | None, str | None, str | None, str | None],
    ] = {}
    for row in cur.fetchall():
        rid = int(row[0])
        out[rid] = (
            _sql_str(row[1]),
            _sql_str(row[2]),
            _sql_str(row[3]),
            _sql_str(row[4]),
            _sql_str(row[5]),
            _sql_str(row[6]) if cod_col else None,
        )
    return out


def _sql_optional_int_positive(val: Any) -> int | None:
    """Entero > 0 desde fila ODBC; None si vacío o cero."""
    if val is None:
        return None
    try:
        f = float(val)
        if f != f:  # NaN
            return None
        i = int(f)
        return i if i > 0 else None
    except (TypeError, ValueError):
        return None


def cargar_indice_articulos(
    cur,
) -> tuple[
    dict[str, str],
    dict[str, tuple[str, str]],
    dict[str, tuple[str | None, int | None, int | None, int | None]],
]:
    """
    Por COD_ARTICULO normalizado y por clave numérica canónica (format).
    El tercer mapa guarda datos de empaque/venta (UNIDAD_MEDIDA, UNID_BULTO, …)
    para interpretar la cantidad del export frente a MSSQL.
    """
    cur.execute(
        """
        SELECT LTRIM(RTRIM(COD_ARTICULO)), LTRIM(RTRIM(ISNULL(DESCRIPCION, ''))),
               LTRIM(RTRIM(ISNULL(UNIDAD_MEDIDA, ''))),
               UNID_BULTO, UNIDAD_MIN_VTA, CTD_UNIDADES
        FROM ARTICULOS
        WHERE COD_ARTICULO IS NOT NULL AND LTRIM(RTRIM(COD_ARTICULO)) <> ''
        """
    )
    by_str: dict[str, str] = {}
    by_num_key: dict[str, tuple[str, str]] = {}
    extras: dict[str, tuple[str | None, int | None, int | None, int | None]] = {}
    for row in cur.fetchall():
        cod_raw, desc, um_raw = row[0], row[1], row[2]
        ub, umv, ctd = row[3], row[4], row[5]
        cod_s = normalize_code(cod_raw)
        if not cod_s:
            continue
        desc = desc or ""
        um = (um_raw or "").strip() or None
        ub_i = _sql_optional_int_positive(ub)
        umv_i = _sql_optional_int_positive(umv)
        ctd_i = _sql_optional_int_positive(ctd)
        by_str.setdefault(cod_s, desc)
        if cod_s not in extras:
            extras[cod_s] = (um, ub_i, umv_i, ctd_i)
        try:
            d = Decimal(cod_s.replace(",", "."))
            nk = format(d.normalize(), "f")
            if nk not in by_num_key:
                by_num_key[nk] = (cod_s, desc)
        except InvalidOperation:
            pass
    return by_str, by_num_key, extras


def lookup_extra_cantidad_mssql(
    extras: dict[str, tuple[str | None, int | None, int | None, int | None]],
    cod_mssql: str | None,
) -> tuple[str | None, int | None, int | None, int | None]:
    if not cod_mssql or not extras:
        return (None, None, None, None)
    return extras.get(normalize_code(cod_mssql), (None, None, None, None))


def armar_contexto_cantidad_mssql(
    unidad_medida: str | None,
    unid_bulto: int | None,
    unidad_min_vta: int | None,
    ctd_unidades: int | None,
) -> str | None:
    """
    Texto orientativo (no sustituye regla de negocio formal). Ver MAPEO_PREVENTAS.
    """
    frases: list[str] = []
    um = (unidad_medida or "").strip().upper()
    if um.startswith("UNI"):
        frases.append(
            "En Gestion, UNIDAD_MEDIDA tipo UNI: la cantidad del preventa suele "
            "interpretarse en unidades de venta (piezas al consumidor), no en bultos."
        )
    elif um.startswith("DIS"):
        frases.append(
            "UNIDAD_MEDIDA tipo DIS (display/exhibidor): conviene cruzar con catálogo; "
            "la cantidad podría ser por display y no por unidad suelta."
        )
    elif um:
        frases.append(f"UNIDAD_MEDIDA en MSSQL: «{unidad_medida.strip()}».")
    if unid_bulto is not None:
        frases.append(
            f"UNID_BULTO={unid_bulto} es dato de empaque/logística en ARTICULOS; "
            "no implica que el CSV esté expresado en ese múltiplo."
        )
    if unidad_min_vta is not None:
        frases.append(
            f"UNIDAD_MIN_VTA={unidad_min_vta}: venta mínima declarada para el código."
        )
    if ctd_unidades is not None:
        frases.append(f"CTD_UNIDADES={ctd_unidades} (referencia en ficha de artículo).")
    if not frases:
        return None
    return " ".join(frases)


def resolver_articulo(
    cod_odoo: str,
    by_str: dict[str, str],
    by_num_key: dict[str, tuple[str, str]],
) -> tuple[str | None, str | None, str]:
    """
    Devuelve (cod_mssql, descripcion, estado).
    estado: ok | sin_codigo | no_en_mssql | ambiguo (no implementado)
    """
    if not cod_odoo:
        return None, None, "sin_codigo"
    c = normalize_code(cod_odoo)
    if c in by_str:
        return c, by_str[c], "ok"
    if c in by_num_key:
        real, desc = by_num_key[c]
        return real, desc, "ok"
    # variante con espacio inicial (Odoo a veces lo guarda así)
    sp = f" {c}"
    if sp in by_str:
        return sp.strip(), by_str[sp], "ok"
    return None, None, "no_en_mssql"


def aplicar_schema_sqlite(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        PRAGMA foreign_keys = ON;
        PRAGMA journal_mode = DELETE;

        CREATE TABLE IF NOT EXISTS meta_carga (
          id INTEGER PRIMARY KEY CHECK (id = 1),
          archivo_origen TEXT,
          cargado_en TEXT NOT NULL,
          filas INTEGER NOT NULL DEFAULT 0,
          odoo_resuelto_en TEXT
        );
        INSERT OR IGNORE INTO meta_carga (id, archivo_origen, cargado_en, filas)
        VALUES (1, NULL, '', 0);

        CREATE TABLE IF NOT EXISTS pedido_lineas (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          operacion TEXT,
          id_vendedor_mssql INTEGER,
          vendedor_nombre TEXT,
          vendedor_codigo TEXT,
          id_cliente_mssql INTEGER,
          cliente_razon_social TEXT,
          cliente_codigo TEXT,
          cliente_nombre_fantasia TEXT,
          cliente_cuit TEXT,
          cliente_direccion TEXT,
          cliente_email TEXT,
          cliente_display TEXT,
          codigo_articulo_raw TEXT,
          codigo_articulo_odoo TEXT,
          cod_articulo_mssql TEXT,
          descripcion_articulo TEXT,
          estado_articulo TEXT,
          cantidad_pedida REAL,
          ruta_raw TEXT,
          fecha_pedido TEXT,
          hora_pedido TEXT,
          date_order_odoo TEXT,
          user_id_odoo INTEGER,
          user_name_odoo TEXT,
          partner_id_odoo INTEGER,
          partner_name_odoo TEXT,
          partner_ref_odoo TEXT,
          product_id_odoo INTEGER,
          product_name_odoo TEXT,
          product_default_code_resuelto TEXT,
          estado_linea_odoo TEXT,
          codigo_odoo_antes_correccion TEXT,
          correccion_codigo_detalle TEXT,
          alerta_articulo_mssql TEXT,
          mssql_unidad_medida TEXT,
          mssql_unid_bulto INTEGER,
          mssql_unidad_min_vta INTEGER,
          mssql_ctd_unidades INTEGER,
          cantidad_pedida_contexto TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_pedido_operacion ON pedido_lineas(operacion);
        CREATE INDEX IF NOT EXISTS idx_pedido_cliente ON pedido_lineas(id_cliente_mssql);
        """
    )
    _ensure_columnas_cliente_mssql(conn)
    _ensure_columnas_codigo_auditoria(conn)
    _ensure_columnas_cantidad_mssql(conn)
    conn.commit()


def _ensure_columnas_codigo_auditoria(conn: sqlite3.Connection) -> None:
    cols = {row[1] for row in conn.execute("PRAGMA table_info(pedido_lineas)")}
    for name, typ in (
        ("codigo_odoo_antes_correccion", "TEXT"),
        ("correccion_codigo_detalle", "TEXT"),
        ("alerta_articulo_mssql", "TEXT"),
    ):
        if name not in cols:
            conn.execute(f"ALTER TABLE pedido_lineas ADD COLUMN {name} {typ}")


def _ensure_columnas_cantidad_mssql(conn: sqlite3.Connection) -> None:
    cols = {row[1] for row in conn.execute("PRAGMA table_info(pedido_lineas)")}
    for name, typ in (
        ("mssql_unidad_medida", "TEXT"),
        ("mssql_unid_bulto", "INTEGER"),
        ("mssql_unidad_min_vta", "INTEGER"),
        ("mssql_ctd_unidades", "INTEGER"),
        ("cantidad_pedida_contexto", "TEXT"),
    ):
        if name not in cols:
            conn.execute(f"ALTER TABLE pedido_lineas ADD COLUMN {name} {typ}")


def _ensure_columnas_cliente_mssql(conn: sqlite3.Connection) -> None:
    """SQLite viejas sin campos extendidos de CLIENTES."""
    cols = {row[1] for row in conn.execute("PRAGMA table_info(pedido_lineas)")}
    for name, typ in (
        ("cliente_nombre_fantasia", "TEXT"),
        ("cliente_cuit", "TEXT"),
        ("cliente_direccion", "TEXT"),
        ("cliente_email", "TEXT"),
        ("cliente_display", "TEXT"),
    ):
        if name not in cols:
            conn.execute(f"ALTER TABLE pedido_lineas ADD COLUMN {name} {typ}")


def leer_filas_csv(path: Path, *, sin_cabecera: bool = False) -> list[dict[str, str]]:
    """
    No usar solo DictReader: muchos exports repiten cabeceras vacías y Python
    deja una sola clave '' (se pierden fecha/hora y columnas intermedias).

    sin_cabecera: el archivo empieza directo en datos (sin fila Operacion,Vendedor,...).
    """
    rows: list[dict[str, str]] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        row_iter = iter(reader)
        if not sin_cabecera:
            try:
                next(row_iter)  # descartar cabecera
            except StopIteration:
                return []
        for parts in row_iter:
            if not parts or not any((p or "").strip() for p in parts):
                continue
            def g(i: int) -> str:
                return (parts[i].strip() if i < len(parts) and parts[i] else "")

            rows.append(
                {
                    "Operacion": g(0),
                    "Vendedor": g(1),
                    "Cliente": g(2),
                    "Ruta": g(3),
                    "Codigo articulo": g(4),
                    "Cantidad pedida": g(5),
                    "Col6": g(6),
                    "Fecha pedido": g(7),
                    "Hora pedido": g(8),
                }
            )
    return rows


def leer_filas_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise SystemExit("Para XLSX: pip install openpyxl") from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return []
    keys = [
        (str(h).strip() if h is not None else "") or f"_col{i}"
        for i, h in enumerate(header)
    ]
    rows: list[dict[str, str]] = []
    for tup in it:
        if tup is None or all(v is None or str(v).strip() == "" for v in tup):
            continue
        d: dict[str, str] = {}
        for i, k in enumerate(keys):
            v = tup[i] if i < len(tup) else None
            if v is None:
                d[k] = ""
            elif isinstance(v, float) and v == int(v):
                d[k] = str(int(v))
            else:
                d[k] = str(v).strip()
        rows.append(d)
    wb.close()
    return rows


def campo(row: dict[str, str], *nombres: str) -> str:
    for n in nombres:
        if n in row and row[n]:
            return row[n].strip()
    return ""


def main() -> int:
    ap = argparse.ArgumentParser(
        description="CSV/XLSX pedido → SQLite enriquecida con MSSQL Gestion"
    )
    ap.add_argument("--csv", type=Path, help="Ruta al CSV")
    ap.add_argument("--xlsx", type=Path, help="Ruta al XLSX (alternativa a --csv)")
    ap.add_argument(
        "--db",
        type=Path,
        default=SCRIPT_DIR / "reportes_pedido_sqlite" / "pedido_enriquecido.sqlite",
        help="Ruta base SQLite de salida",
    )
    ap.add_argument(
        "--export-csv",
        type=Path,
        default=None,
        help="Si se indica, escribe CSV enriquecido (UTF-8)",
    )
    ap.add_argument(
        "--sin-mssql",
        action="store_true",
        help="No conecta a MSSQL: solo vuelca líneas con códigos Odoo; "
        "nombres y descripciones quedan vacíos (estado_articulo=sin_mssql)",
    )
    ap.add_argument(
        "--resolver-odoo",
        action="store_true",
        help="Tras cargar el CSV, consulta Odoo (master_18) y rellena ids/nombres "
        "en columnas *_odoo (requiere JSON de mapeo y red)",
    )
    ap.add_argument(
        "--mapeo",
        type=Path,
        default=SCRIPT_DIR / "mapeo_preventas_master18.json",
        help="JSON vendedores/clientes (con --resolver-odoo)",
    )
    ap.add_argument(
        "--append",
        action="store_true",
        help="Agregar líneas al .sqlite existente (no borra la base). Actualiza meta_carga.",
    )
    ap.add_argument(
        "--sin-cabecera-csv",
        action="store_true",
        help="El CSV no trae fila de títulos (primera fila ya es un pedido).",
    )
    ap.add_argument(
        "--correcciones",
        type=Path,
        default=DEFAULT_CORRECCIONES_JSON,
        help="JSON raw_a_cod_odoo / codigo_odoo_a_correcto (ver .example.json)",
    )
    args = ap.parse_args()

    if args.resolver_odoo and not args.mapeo.is_file():
        raise SystemExit(
            f"--resolver-odoo requiere el JSON de mapeo (no existe): {args.mapeo}"
        )

    if bool(args.csv) == bool(args.xlsx):
        raise SystemExit("Indica exactamente uno de: --csv o --xlsx")

    src = args.csv or args.xlsx
    if not src.is_file():
        raise SystemExit(f"No existe el archivo: {src}")

    if args.csv:
        filas = leer_filas_csv(args.csv, sin_cabecera=args.sin_cabecera_csv)
    else:
        filas = leer_filas_xlsx(args.xlsx)

    ids_v: set[int] = set()
    ids_c: set[int] = set()
    for r in filas:
        v = campo(r, "Vendedor")
        c = campo(r, "Cliente")
        if v.isdigit():
            ids_v.add(int(v))
        if c.isdigit():
            ids_c.add(int(c))

    args.db.parent.mkdir(parents=True, exist_ok=True)
    if args.append:
        if not args.db.is_file():
            raise SystemExit(
                f"--append requiere una base existente: {args.db}"
            )
    elif args.db.is_file():
        args.db.unlink()

    if args.correcciones.is_file():
        corr_data = cargar_correcciones(args.correcciones)
        nreg = len(corr_data["raw_a_cod_odoo"]) + len(
            corr_data["codigo_odoo_a_correcto"]
        )
        print(
            f"Correcciones código: {nreg} reglas desde {args.correcciones.name}"
        )
    else:
        if args.correcciones != DEFAULT_CORRECCIONES_JSON:
            print(
                f"Aviso: no existe {args.correcciones}, se omite mapa de correcciones"
            )
        corr_data = cargar_correcciones(None)

    sl = sqlite3.connect(args.db)
    sl.row_factory = sqlite3.Row
    aplicar_schema_sqlite(sl)

    prev_origen = ""
    if args.append:
        row_m = sl.execute(
            "SELECT archivo_origen FROM meta_carga WHERE id = 1"
        ).fetchone()
        if row_m:
            prev_origen = (row_m[0] or "").strip()

    if args.sin_mssql:
        mv, mc = {}, {}
        by_str, by_num, extras_cant = {}, {}, {}
    else:
        mssql = conectar_mssql()
        try:
            cur = mssql.cursor()
            cod_col = detect_cliente_cod_column(cur)
            v_nom, v_cod = detect_vendedor_columns(cur)
            mv = cargar_mapa_vendedores(cur, ids_v, v_nom, v_cod)
            mc = cargar_mapa_clientes(cur, ids_c, cod_col)
            by_str, by_num, extras_cant = cargar_indice_articulos(cur)
        finally:
            mssql.close()

    now = datetime.now().isoformat(timespec="seconds")
    n = 0
    for r in filas:
        op = campo(r, "Operacion", "Numero de Operacion")
        vend = campo(r, "Vendedor")
        cli = campo(r, "Cliente")
        cod_raw = campo(r, "Codigo articulo")
        qty_s = campo(r, "Cantidad pedida")
        ruta = campo(r, "Ruta")
        fecha_p = campo(r, "Fecha pedido")
        hora_p = campo(r, "Hora pedido")

        id_v = int(vend) if vend.isdigit() else None
        id_c = int(cli) if cli.isdigit() else None
        vn, vc = mv.get(id_v, (None, None)) if id_v is not None else (None, None)
        if id_c is not None:
            crz, cnf, ccuit, cdir, cmail, ccod = mc.get(id_c, (None,) * 6)
            cdis = armar_cliente_display(crz, cnf)
        else:
            crz = cnf = ccuit = cdir = cmail = ccod = cdis = None

        cod_calc = articulo_csv_a_default_code(cod_raw)
        cod_odoo, odoo_prev_corr, corr_detalle = aplicar_correcciones_codigo(
            cod_raw, cod_calc, corr_data
        )
        if args.sin_mssql:
            ca_mssql, desc_a, est_a = None, None, "sin_mssql"
            alerta_ms = None
        else:
            ca_mssql, desc_a, est_a = resolver_articulo(cod_odoo, by_str, by_num)
            alerta_ms = alerta_por_descripcion_mssql(desc_a)

        um_ex, ub_ex, umv_ex, ctd_ex = lookup_extra_cantidad_mssql(
            extras_cant, ca_mssql
        )
        ctx_qty = (
            armar_contexto_cantidad_mssql(um_ex, ub_ex, umv_ex, ctd_ex)
            if not args.sin_mssql
            else None
        )

        try:
            qty = float(qty_s.replace(",", ".")) if qty_s else None
        except ValueError:
            qty = None

        sl.execute(
            """
            INSERT INTO pedido_lineas (
              operacion, id_vendedor_mssql, vendedor_nombre, vendedor_codigo,
              id_cliente_mssql, cliente_razon_social, cliente_codigo,
              cliente_nombre_fantasia, cliente_cuit, cliente_direccion, cliente_email,
              cliente_display,
              codigo_articulo_raw, codigo_articulo_odoo, cod_articulo_mssql,
              descripcion_articulo, estado_articulo, cantidad_pedida,
              ruta_raw, fecha_pedido, hora_pedido,
              codigo_odoo_antes_correccion, correccion_codigo_detalle, alerta_articulo_mssql,
              mssql_unidad_medida, mssql_unid_bulto, mssql_unidad_min_vta,
              mssql_ctd_unidades, cantidad_pedida_contexto
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                op,
                id_v,
                vn,
                vc,
                id_c,
                crz,
                ccod,
                cnf,
                ccuit,
                cdir,
                cmail,
                cdis,
                cod_raw,
                cod_odoo,
                ca_mssql,
                desc_a,
                est_a,
                qty,
                ruta or None,
                fecha_p or None,
                hora_p or None,
                odoo_prev_corr,
                corr_detalle,
                alerta_ms,
                um_ex,
                ub_ex,
                umv_ex,
                ctd_ex,
                ctx_qty,
            ),
        )
        n += 1

    total_filas = sl.execute("SELECT COUNT(*) FROM pedido_lineas").fetchone()[0]
    if args.append and prev_origen:
        origen_meta = f"{prev_origen} | + {src}"
    else:
        origen_meta = str(src)
    sl.execute(
        "UPDATE meta_carga SET archivo_origen = ?, cargado_en = ?, filas = ? WHERE id = 1",
        (origen_meta, now, total_filas),
    )
    n_corr = sl.execute(
        "SELECT COUNT(*) FROM pedido_lineas WHERE correccion_codigo_detalle IS NOT NULL"
    ).fetchone()[0]
    n_zz = sl.execute(
        "SELECT COUNT(*) FROM pedido_lineas WHERE alerta_articulo_mssql IS NOT NULL"
    ).fetchone()[0]
    sl.commit()

    if n_corr:
        print(f"Líneas con corrección de código: {n_corr}")
    if n_zz:
        print(f"Líneas con alerta ZZ/ZZZ en descripción MSSQL: {n_zz}")

    if args.resolver_odoo:
        from pedido_sqlite_odoo import resolver_pedido_sqlite_odoo

        st = resolver_pedido_sqlite_odoo(sl, args.mapeo)
        print(
            "Odoo: líneas resueltas "
            f"(ok={st['linea_ok']}, con incidencias={st['con_algun_fallo']}, total={st['filas']})"
        )

    sl.close()

    print(f"SQLite: {args.db}")
    print(f"Filas insertadas (este CSV): {n}")
    print(f"Filas totales en pedido_lineas: {total_filas}")

    if args.export_csv:
        import csv as csv_mod

        sl = sqlite3.connect(args.db)
        sl.row_factory = sqlite3.Row
        cur2 = sl.execute(
            "SELECT * FROM pedido_lineas ORDER BY operacion, id"
        )
        args.export_csv.parent.mkdir(parents=True, exist_ok=True)
        cols = [d[0] for d in cur2.description]
        with open(args.export_csv, "w", newline="", encoding="utf-8") as f:
            w = csv_mod.writer(f)
            w.writerow(cols)
            for row in cur2:
                w.writerow(list(row))
        sl.close()
        print(f"CSV export: {args.export_csv}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
