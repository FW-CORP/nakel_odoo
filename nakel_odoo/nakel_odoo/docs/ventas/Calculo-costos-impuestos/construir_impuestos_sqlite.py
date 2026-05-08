#!/usr/bin/env python3
"""
Arma ``impuestos.sqlite`` a partir de:
- ``Impuestos Internos completo.xlsx`` (columna A = códigos a incluir)
- ``productos_stock_nakel.sqlite`` (referencia_interna, nombre, plu)
- Opcional: ``Stock valuado a precio de costo.xls`` (OOXML/ZIP). Detecta columnas por
  cabecera: export **corto** (A código, B descripción, C costo sin IVA) o **largo**
  con CxB y costo/imp./neto en D–F.

Fórmula planilla II (col G): ``costo_sin_iva_sin_imp_interno = D / (E + 1)`` donde **D** es
``costo_sin_iva`` (prioridad: valor ya cargado desde stock valuado; si falta, col. D del Excel II)
y **E** es ``impuesto_interno`` (prioridad: col. E del Excel II; si falta, valor ya en DB).

Uso:
  python3 construir_impuestos_sqlite.py
  python3 construir_impuestos_sqlite.py --sin-stock-valuado
  python3 construir_impuestos_sqlite.py --solo-actualizar-stock-valuado
  python3 construir_impuestos_sqlite.py --solo-aplicar-formula-g
  python3 construir_impuestos_sqlite.py --solo-cargar-precios-venta --csv-precios-venta precios_venta.csv

Precios de venta: columnas ``precio_venta_sin_iva`` y ``precio_venta_con_iva``. Se pueden cargar desde
columnas opcionales en el Excel II (cabeceras con «venta» + «sin»/«con» + «iva») y/o ``--csv-precios-venta``.
"""
from __future__ import annotations

import argparse
import csv
import io
import sqlite3
import sys
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    print("Falta openpyxl. Ej.: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1) from e

_REPO_ROOT = Path(__file__).resolve().parents[2]
_DV = _REPO_ROOT / "db_precios_ventas"
if str(_DV) not in sys.path:
    sys.path.insert(0, str(_DV))

from codigo_normalizar import normalize_code  # noqa: E402

_DIR = Path(__file__).resolve().parent
_DEFAULT_EXCEL = _DIR / "Impuestos Internos completo.xlsx"
_DEFAULT_SALIDA = _DIR / "impuestos.sqlite"
_DEFAULT_STOCK = Path(
    "/home/klap/Descargas/nakel_tempo_abril/canonicos_marcelo/productos_stock_nakel.sqlite"
)
_DEFAULT_STOCK_VALUADO = _DIR / "Stock valuado a precio de costo.xls"
_DEFAULT_CSV_PRECIOS_VENTA = _DIR / "precios_venta.csv"


def _open_workbook_readonly(path: Path):
    """``.xls`` que es OOXML (ZIP) o ``.xlsx``."""
    data = path.read_bytes()
    if len(data) >= 2 and data[:2] == b"PK":
        return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    suf = path.suffix.lower()
    if suf in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    raise SystemExit(
        f"{path.name}: .xls binario (BIFF) no soportado; exportar como xlsx OOXML."
    )


def _num_a_float(val: object) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(Decimal(str(val).replace(",", ".")))
    except (InvalidOperation, ValueError):
        return None


def codigos_desde_excel(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    orden: list[str] = []
    vistos: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        raw = row[0]
        if raw is None:
            continue
        c = normalize_code(raw)
        if not c:
            continue
        if c not in vistos:
            vistos.add(c)
            orden.append(c)
    wb.close()
    return orden


def leer_ii_d_e_g(path: Path) -> dict[str, tuple[float | None, float | None, float | None]]:
    """
    Por código normalizado: (D Costo sin IVA, E Imp. interno., G Costo sin Imp. Interno si existe en hoja).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    out: dict[str, tuple[float | None, float | None, float | None]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        code = normalize_code(row[0])
        if not code:
            continue
        d = _num_a_float(row[3]) if len(row) > 3 else None
        e = _num_a_float(row[4]) if len(row) > 4 else None
        g = _num_a_float(row[6]) if len(row) > 6 else None
        out[code] = (d, e, g)
    wb.close()
    return out


def ensure_columnas_precio_venta(cur: sqlite3.Cursor) -> None:
    cols = {r[1] for r in cur.execute("PRAGMA table_info(productos)")}
    if "precio_venta_sin_iva" not in cols:
        cur.execute("ALTER TABLE productos ADD COLUMN precio_venta_sin_iva REAL")
    if "precio_venta_con_iva" not in cols:
        cur.execute("ALTER TABLE productos ADD COLUMN precio_venta_con_iva REAL")


def leer_precios_venta_desde_excel(path: Path) -> dict[str, tuple[float | None, float | None]]:
    """
    Detecta columnas cuya cabecera contiene «venta» y («sin»+«iva») o («con»+«iva»).
    No confunde con «costo sin IVA» (no lleva «venta»).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    labels = [str(x or "").strip().lower().replace("í", "i") for x in header]
    i_sin = i_con = None
    for i, s in enumerate(labels):
        if not s or "venta" not in s:
            continue
        if "sin" in s and "iva" in s:
            i_sin = i
        elif "con" in s and "iva" in s:
            i_con = i
    if i_sin is None and i_con is None:
        wb.close()
        return {}
    out: dict[str, tuple[float | None, float | None]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        code = normalize_code(row[0])
        if not code:
            continue
        ps = _num_a_float(row[i_sin]) if i_sin is not None and len(row) > i_sin else None
        pc = _num_a_float(row[i_con]) if i_con is not None and len(row) > i_con else None
        if ps is None and pc is None:
            continue
        out[code] = (ps, pc)
    wb.close()
    return out


def leer_precios_venta_csv(path: Path) -> dict[str, tuple[float | None, float | None]]:
    """CSV con cabeceras: referencia_interna/codigo; precio_venta_sin_iva/venta_sin_iva; precio_venta_con_iva/venta_con_iva."""
    out: dict[str, tuple[float | None, float | None]] = {}
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return out
        fn = {h.strip().lower(): h for h in reader.fieldnames if h}

        def pick(*names: str) -> str | None:
            for n in names:
                if n in fn:
                    return fn[n]
            return None

        k_ref = pick("referencia_interna", "codigo", "ref", "codigo_interno")
        k_sin = pick(
            "precio_venta_sin_iva",
            "venta_sin_iva",
            "pventa_sin_iva",
            "precio_sin_iva_venta",
        )
        k_con = pick(
            "precio_venta_con_iva",
            "venta_con_iva",
            "pventa_con_iva",
            "precio_con_iva_venta",
        )
        if not k_ref:
            return out
        for row in reader:
            raw = row.get(k_ref)
            if raw is None or str(raw).strip() == "":
                continue
            code = normalize_code(raw)
            if not code:
                continue
            ps = _num_a_float(row[k_sin]) if k_sin else None
            pc = _num_a_float(row[k_con]) if k_con else None
            if ps is None and pc is None:
                continue
            out[code] = (ps, pc)
    return out


def aplicar_precios_venta(
    con: sqlite3.Connection, precios: dict[str, tuple[float | None, float | None]]
) -> dict[str, int]:
    """Actualiza solo valores no NULL del mapa."""
    cur = con.cursor()
    ensure_columnas_precio_venta(cur)
    ok = 0
    sin_fila = 0
    for code, (ps, pc) in precios.items():
        ref_row = cur.execute(
            "SELECT referencia_interna FROM productos WHERE referencia_interna = ?",
            (code,),
        ).fetchone()
        if ref_row is None:
            for r in cur.execute("SELECT referencia_interna FROM productos"):
                if normalize_code(r[0]) == code:
                    ref_row = r
                    break
        if ref_row is None:
            sin_fila += 1
            continue
        ref = ref_row[0]
        parts: list[str] = []
        vals: list[float | str] = []
        if ps is not None:
            parts.append("precio_venta_sin_iva = ?")
            vals.append(ps)
        if pc is not None:
            parts.append("precio_venta_con_iva = ?")
            vals.append(pc)
        if not parts:
            continue
        vals.append(ref)
        cur.execute(
            f"UPDATE productos SET {', '.join(parts)} WHERE referencia_interna = ?",
            vals,
        )
        ok += 1
    return {"actualizados": ok, "sin_producto": sin_fila}


def aplicar_formula_g_costo_neto(
    con: sqlite3.Connection, ii_map: dict[str, tuple[float | None, float | None, float | None]]
) -> dict[str, int]:
    """
    ``costo_sin_iva_sin_imp_interno = D / (E + 1)``; actualiza ``impuesto_interno`` con E usado.
    """
    cur = con.cursor()
    ensure_columnas_precio_venta(cur)
    referencias = [r[0] for r in cur.execute("SELECT referencia_interna FROM productos")]
    ok = 0
    sin_d = 0
    sin_e = 0
    sin_map = 0
    for ref in referencias:
        key = normalize_code(ref)
        tri = ii_map.get(key)
        if tri is None:
            sin_map += 1
            continue
        d_ii, e_ii, _g_xls = tri
        row = cur.execute(
            "SELECT costo_sin_iva, impuesto_interno FROM productos WHERE referencia_interna = ?",
            (ref,),
        ).fetchone()
        d_db, imp_db = row[0], row[1]
        d = d_db if d_db is not None else d_ii
        e = e_ii if e_ii is not None else imp_db
        if d is None:
            sin_d += 1
            continue
        if e is None:
            sin_e += 1
            continue
        f = e + 1.0
        if abs(f) < 1e-15:
            continue
        g = d / f
        cur.execute(
            """
            UPDATE productos
            SET impuesto_interno = ?, costo_sin_iva_sin_imp_interno = ?
            WHERE referencia_interna = ?
            """,
            (e, g, ref),
        )
        ok += 1
    return {"actualizados": ok, "sin_d": sin_d, "sin_e": sin_e, "sin_fila_ii": sin_map}


def _norm_hdr_stock(c: object) -> str:
    if c is None:
        return ""
    return str(c).strip().lower().replace("í", "i")


def _indices_stock_valuado(header_row: tuple) -> tuple[int | None, int | None, int | None]:
    """
    Índices 0-based: costo_sin_iva, impuesto_interno, costo_sin_iva_sin_imp_interno.
    Soporta export con solo 3 columnas de dato (costo en C) o planilla extendida con CxB.
    """
    idx_costo = idx_ii = idx_neto = None
    cells = list(enumerate(header_row or ()))
    for i, cell in cells:
        s = _norm_hdr_stock(cell)
        if not s:
            continue
        if "costo" in s and "iva" in s and ("y sin imp" in s or "sin imp. interno" in s):
            idx_neto = i
    for i, cell in cells:
        s = _norm_hdr_stock(cell)
        if not s or i == idx_neto:
            continue
        if "imp" in s and "intern" in s:
            idx_ii = i
    for i, cell in cells:
        s = _norm_hdr_stock(cell)
        if not s or i in (idx_neto, idx_ii):
            continue
        if "costo" in s and "iva" in s:
            idx_costo = i
    if idx_costo is None:
        h2 = [_norm_hdr_stock(c) for c in (header_row or ())]
        if len(h2) >= 6 and h2[2] == "cxb":
            idx_costo, idx_ii, idx_neto = 3, 4, 5
        elif len(h2) >= 3 and h2[2] and "costo" in h2[2] and "iva" in h2[2]:
            idx_costo = 2
    return idx_costo, idx_ii, idx_neto


def leer_stock_valuado_costos(path: Path) -> tuple[dict[str, tuple[float | None, float | None, float | None]], int]:
    """
    Lee filas de datos (desde fila 2). Mapea columnas según la fila de cabeceras.
    """
    wb = _open_workbook_readonly(path)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    i_d, i_e, i_f = _indices_stock_valuado(header)
    if i_d is None:
        print(
            "Aviso: Stock valuado sin columna reconocible «Costo sin IVA»; no se cargan costos.",
            file=sys.stderr,
        )
    else:
        print(
            f"Stock valuado: columnas detectadas — costo índice {i_d}, "
            f"imp. interno {i_e}, neto sin II {i_f}",
            file=sys.stderr,
        )

    if i_d is None:
        wb.close()
        return {}, 0

    def _cell(row: tuple, idx: int | None) -> float | None:
        if idx is None or len(row) <= idx:
            return None
        return _num_a_float(row[idx])

    out: dict[str, tuple[float | None, float | None, float | None]] = {}
    duplicados = Counter()
    filas_con_codigo = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        code = normalize_code(row[0])
        if not code:
            continue
        filas_con_codigo += 1
        d = _cell(row, i_d)
        e = _cell(row, i_e)
        fval = _cell(row, i_f)
        if code in out:
            duplicados[code] += 1
        out[code] = (d, e, fval)
    wb.close()
    if duplicados:
        n = sum(duplicados.values())
        print(
            f"Aviso: {n} filas duplican código en stock valuado (última fila gana).",
            file=sys.stderr,
        )
    return out, filas_con_codigo


def _ensure_columnas_costo(cur: sqlite3.Cursor) -> None:
    cols = {r[1] for r in cur.execute("PRAGMA table_info(productos)")}
    for nombre, tipo in (
        ("costo_sin_iva", "REAL"),
        ("impuesto_interno", "REAL"),
        ("costo_sin_iva_sin_imp_interno", "REAL"),
    ):
        if nombre not in cols:
            cur.execute(f"ALTER TABLE productos ADD COLUMN {nombre} {tipo}")


def aplicar_costos_stock_valuado(
    con: sqlite3.Connection, costos: dict[str, tuple[float | None, float | None, float | None]]
) -> dict[str, int]:
    cur = con.cursor()
    ensure_columnas_precio_venta(cur)
    _ensure_columnas_costo(cur)
    cur.execute("UPDATE productos SET costo_sin_iva = NULL, impuesto_interno = NULL, costo_sin_iva_sin_imp_interno = NULL")
    referencias = [r[0] for r in cur.execute("SELECT referencia_interna FROM productos")]
    actualizados = 0
    sin_planilla = 0
    for ref in referencias:
        key = normalize_code(ref)
        trip = costos.get(key)
        if trip is None:
            sin_planilla += 1
            continue
        d, e, fval = trip
        cur.execute(
            """
            UPDATE productos
            SET costo_sin_iva = ?, impuesto_interno = ?, costo_sin_iva_sin_imp_interno = ?
            WHERE referencia_interna = ?
            """,
            (d, e, fval, ref),
        )
        actualizados += 1
    return {"actualizados": actualizados, "sin_fila_stock_valuado": sin_planilla}


def _meta_upsert(cur: sqlite3.Cursor, clave: str, valor: str) -> None:
    cur.execute(
        "INSERT INTO meta (clave, valor) VALUES (?, ?) ON CONFLICT(clave) DO UPDATE SET valor = excluded.valor",
        (clave, valor),
    )


def main() -> None:
    p = argparse.ArgumentParser(description="Crea o actualiza impuestos.sqlite (II + stock valuado).")
    p.add_argument("--excel", type=Path, default=_DEFAULT_EXCEL, help="XLSX impuestos internos (lista de códigos)")
    p.add_argument("--stock", type=Path, default=_DEFAULT_STOCK, help="productos_stock_nakel.sqlite")
    p.add_argument("--salida", type=Path, default=_DEFAULT_SALIDA, help="SQLite de salida")
    p.add_argument(
        "--stock-valuado",
        type=Path,
        default=_DEFAULT_STOCK_VALUADO,
        help="Stock valuado a precio de costo (.xls OOXML o .xlsx)",
    )
    p.add_argument("--sin-stock-valuado", action="store_true", help="No cargar columnas D/E/F")
    p.add_argument(
        "--solo-actualizar-stock-valuado",
        action="store_true",
        help="Solo rellenar D/E/F en la SQLite ya existente (no recrear filas)",
    )
    p.add_argument(
        "--solo-aplicar-formula-g",
        action="store_true",
        help="Solo calcular costo_sin_iva_sin_imp_interno = D/(E+1) desde Excel II + DB",
    )
    p.add_argument(
        "--sin-formula-g",
        action="store_true",
        help="Al crear la base desde cero, no aplicar la fórmula G (solo insert inicial)",
    )
    p.add_argument(
        "--csv-precios-venta",
        type=Path,
        default=None,
        help="CSV con referencia_interna/codigo y precio_venta_sin_iva / precio_venta_con_iva",
    )
    p.add_argument(
        "--sin-cargar-precios-venta",
        action="store_true",
        help="No cargar precios de venta (ni Excel opcional ni CSV)",
    )
    p.add_argument(
        "--sin-precios-venta-csv-auto",
        action="store_true",
        help="No usar automáticamente precios_venta.csv del directorio del script",
    )
    p.add_argument(
        "--solo-cargar-precios-venta",
        action="store_true",
        help="Solo actualizar precio_venta_sin_iva / precio_venta_con_iva desde Excel II y/o CSV",
    )
    args = p.parse_args()

    def _ruta_csv_precios_venta() -> Path | None:
        if args.csv_precios_venta is not None:
            return args.csv_precios_venta
        if not args.sin_precios_venta_csv_auto and _DEFAULT_CSV_PRECIOS_VENTA.is_file():
            return _DEFAULT_CSV_PRECIOS_VENTA
        return None

    def _armar_mapa_precios_venta() -> dict[str, tuple[float | None, float | None]]:
        pv: dict[str, tuple[float | None, float | None]] = {}
        if args.excel.is_file():
            pv.update(leer_precios_venta_desde_excel(args.excel))
        cp = _ruta_csv_precios_venta()
        if cp is not None and cp.is_file():
            pv.update(leer_precios_venta_csv(cp))
        return pv

    if args.solo_cargar_precios_venta:
        if not args.salida.is_file():
            print(f"No existe SQLite: {args.salida}", file=sys.stderr)
            raise SystemExit(1)
        cp = _ruta_csv_precios_venta()
        if not args.excel.is_file() and (cp is None or not cp.is_file()):
            print("Indicá --excel existente o un CSV (--csv-precios-venta o precios_venta.csv).", file=sys.stderr)
            raise SystemExit(1)
        pv = _armar_mapa_precios_venta()
        if not pv:
            print("No se encontraron columnas de venta en el Excel ni filas en el CSV.", file=sys.stderr)
            raise SystemExit(1)
        con = sqlite3.connect(args.salida)
        st = aplicar_precios_venta(con, pv)
        cur = con.cursor()
        _meta_upsert(cur, "fuente_precios_venta", "excel II (si hay columnas) + CSV")
        _meta_upsert(cur, "precios_venta_actualizados", str(st["actualizados"]))
        _meta_upsert(cur, "precios_venta_sin_producto", str(st["sin_producto"]))
        con.commit()
        con.close()
        print(f"SQLite: {args.salida}")
        print(f"Precios venta: actualizados {st['actualizados']} | código sin fila en DB {st['sin_producto']}")
        return

    if args.solo_aplicar_formula_g:
        if not args.salida.is_file():
            print(f"No existe SQLite: {args.salida}", file=sys.stderr)
            raise SystemExit(1)
        if not args.excel.is_file():
            print(f"No existe Excel II: {args.excel}", file=sys.stderr)
            raise SystemExit(1)
        ii_map = leer_ii_d_e_g(args.excel)
        con = sqlite3.connect(args.salida)
        st = aplicar_formula_g_costo_neto(con, ii_map)
        cur = con.cursor()
        _meta_upsert(cur, "formula_g", "costo_sin_iva_sin_imp_interno = D/(E+1); D=COALESCE(costo_sin_iva,II!D); E=COALESCE(II!E,impuesto_interno)")
        _meta_upsert(cur, "formula_g_actualizados", str(st["actualizados"]))
        _meta_upsert(cur, "formula_g_sin_d", str(st["sin_d"]))
        _meta_upsert(cur, "formula_g_sin_e", str(st["sin_e"]))
        _meta_upsert(cur, "formula_g_sin_fila_ii", str(st["sin_fila_ii"]))
        con.commit()
        con.close()
        print(f"SQLite: {args.salida}")
        print(
            f"Fórmula G: actualizados {st['actualizados']} | sin D {st['sin_d']} | sin E {st['sin_e']} | sin fila II {st['sin_fila_ii']}"
        )
        return

    if args.solo_actualizar_stock_valuado:
        if not args.salida.is_file():
            print(f"No existe SQLite: {args.salida}", file=sys.stderr)
            raise SystemExit(1)
        if args.sin_stock_valuado:
            print("Incompatible: --solo-actualizar-stock-valuado requiere planilla stock valuado.", file=sys.stderr)
            raise SystemExit(1)
        if not args.stock_valuado.is_file():
            print(f"No existe planilla stock valuado: {args.stock_valuado}", file=sys.stderr)
            raise SystemExit(1)
        costos, nfilas = leer_stock_valuado_costos(args.stock_valuado)
        con = sqlite3.connect(args.salida)
        stats = aplicar_costos_stock_valuado(con, costos)
        cur = con.cursor()
        _meta_upsert(cur, "fuente_stock_valuado", str(args.stock_valuado.resolve()))
        _meta_upsert(cur, "stock_valuado_filas_con_codigo", str(nfilas))
        _meta_upsert(cur, "stock_valuado_codigos_unicos", str(len(costos)))
        _meta_upsert(cur, "productos_actualizados_stock_valuado", str(stats["actualizados"]))
        _meta_upsert(cur, "productos_sin_fila_stock_valuado", str(stats["sin_fila_stock_valuado"]))
        con.commit()
        con.close()
        print(f"SQLite: {args.salida}")
        print(
            f"Stock valuado: {nfilas} filas con código, {len(costos)} códigos únicos | "
            f"Actualizados en productos: {stats['actualizados']} | Sin fila en planilla: {stats['sin_fila_stock_valuado']}"
        )
        return

    if not args.excel.is_file():
        print(f"No existe Excel: {args.excel}", file=sys.stderr)
        raise SystemExit(1)
    if not args.stock.is_file():
        print(f"No existe SQLite stock: {args.stock}", file=sys.stderr)
        raise SystemExit(1)

    codigos = codigos_desde_excel(args.excel)
    if not codigos:
        print("No hay códigos en columna A del Excel.", file=sys.stderr)
        raise SystemExit(1)

    costos: dict[str, tuple[float | None, float | None, float | None]] = {}
    nfilas_sv = 0
    if not args.sin_stock_valuado and args.stock_valuado.is_file():
        costos, nfilas_sv = leer_stock_valuado_costos(args.stock_valuado)
    elif not args.sin_stock_valuado:
        print(f"Aviso: no se encontró {args.stock_valuado} (columnas D/E/F quedan NULL).", file=sys.stderr)

    con_stock = sqlite3.connect(args.stock)
    con_stock.row_factory = sqlite3.Row
    cur_s = con_stock.cursor()

    if args.salida.exists():
        args.salida.unlink()

    con_out = sqlite3.connect(args.salida)
    cur_o = con_out.cursor()
    st_pv = None
    cur_o.execute(
        """
        CREATE TABLE productos (
            referencia_interna TEXT PRIMARY KEY NOT NULL,
            nombre TEXT,
            plu TEXT,
            costo_sin_iva REAL,
            impuesto_interno REAL,
            costo_sin_iva_sin_imp_interno REAL,
            precio_venta_sin_iva REAL,
            precio_venta_con_iva REAL
        )
        """
    )
    cur_o.execute(
        """
        CREATE TABLE meta (
            clave TEXT PRIMARY KEY,
            valor TEXT
        )
        """
    )

    insertados = 0
    faltantes: list[str] = []
    for code in codigos:
        cur_s.execute(
            "SELECT referencia_interna, nombre, plu FROM productos WHERE referencia_interna = ?",
            (code,),
        )
        row = cur_s.fetchone()
        if row is None:
            faltantes.append(code)
            continue
        ref = row["referencia_interna"]
        trip = costos.get(normalize_code(ref))
        cur_o.execute(
            """
            INSERT INTO productos (
                referencia_interna, nombre, plu,
                costo_sin_iva, impuesto_interno, costo_sin_iva_sin_imp_interno,
                precio_venta_sin_iva, precio_venta_con_iva
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                ref,
                row["nombre"],
                row["plu"],
                trip[0] if trip else None,
                trip[1] if trip else None,
                trip[2] if trip else None,
                None,
                None,
            ),
        )
        insertados += 1

    cur_o.execute("INSERT INTO meta (clave, valor) VALUES (?, ?)", ("fuente_excel", str(args.excel.resolve())))
    cur_o.execute(
        "INSERT INTO meta (clave, valor) VALUES (?, ?)",
        ("fuente_stock_sqlite", str(args.stock.resolve())),
    )
    cur_o.execute("INSERT INTO meta (clave, valor) VALUES (?, ?)", ("codigos_en_excel", str(len(codigos))))
    cur_o.execute("INSERT INTO meta (clave, valor) VALUES (?, ?)", ("filas_insertadas", str(insertados)))
    if faltantes:
        cur_o.execute(
            "INSERT INTO meta (clave, valor) VALUES (?, ?)",
            ("codigos_excel_sin_match_stock", ",".join(faltantes)),
        )
    if costos:
        cur_o.execute(
            "INSERT INTO meta (clave, valor) VALUES (?, ?)",
            ("fuente_stock_valuado", str(args.stock_valuado.resolve())),
        )
        cur_o.execute(
            "INSERT INTO meta (clave, valor) VALUES (?, ?)",
            ("stock_valuado_filas_con_codigo", str(nfilas_sv)),
        )
        cur_o.execute(
            "INSERT INTO meta (clave, valor) VALUES (?, ?)",
            ("stock_valuado_codigos_unicos", str(len(costos))),
        )
        nc = cur_o.execute(
            "SELECT COUNT(*) FROM productos WHERE costo_sin_iva IS NOT NULL"
        ).fetchone()[0]
        cur_o.execute(
            "INSERT INTO meta (clave, valor) VALUES (?, ?)",
            ("productos_con_costo_stock_valuado", str(nc)),
        )

    if not args.sin_formula_g:
        ii_map = leer_ii_d_e_g(args.excel)
        st_g = aplicar_formula_g_costo_neto(con_out, ii_map)
        cur_o.execute(
            "INSERT INTO meta (clave, valor) VALUES (?, ?)",
            (
                "formula_g",
                "costo_sin_iva_sin_imp_interno = D/(E+1); D=COALESCE(costo_sin_iva,II!D); E=COALESCE(II!E,impuesto_interno)",
            ),
        )
        cur_o.execute(
            "INSERT INTO meta (clave, valor) VALUES (?, ?)",
            ("formula_g_actualizados", str(st_g["actualizados"])),
        )
        cur_o.execute(
            "INSERT INTO meta (clave, valor) VALUES (?, ?)",
            ("formula_g_sin_d", str(st_g["sin_d"])),
        )
        cur_o.execute(
            "INSERT INTO meta (clave, valor) VALUES (?, ?)",
            ("formula_g_sin_e", str(st_g["sin_e"])),
        )

    if not args.sin_cargar_precios_venta:
        pv_map = _armar_mapa_precios_venta()
        if pv_map:
            ensure_columnas_precio_venta(cur_o)
            st_pv = aplicar_precios_venta(con_out, pv_map)
            cur_o.execute(
                "INSERT INTO meta (clave, valor) VALUES (?, ?)",
                ("precios_venta_actualizados", str(st_pv["actualizados"])),
            )
            cur_o.execute(
                "INSERT INTO meta (clave, valor) VALUES (?, ?)",
                ("precios_venta_sin_producto_en_db", str(st_pv["sin_producto"])),
            )
            csv_p = _ruta_csv_precios_venta()
            if csv_p is not None and csv_p.is_file():
                cur_o.execute(
                    "INSERT INTO meta (clave, valor) VALUES (?, ?)",
                    ("fuente_precios_venta_csv", str(csv_p.resolve())),
                )

    con_out.commit()
    con_out.close()
    con_stock.close()

    print(f"SQLite: {args.salida}")
    print(f"Códigos en Excel: {len(codigos)} | Insertados: {insertados}")
    if costos:
        n_cc = sqlite3.connect(args.salida)
        nc = n_cc.execute(
            "SELECT COUNT(*) FROM productos WHERE costo_sin_iva IS NOT NULL"
        ).fetchone()[0]
        n_cc.close()
        print(
            f"Stock valuado: {nfilas_sv} filas, {len(costos)} códigos únicos | "
            f"Productos con col. D no NULL: {nc}"
        )
    if faltantes:
        print(f"Sin match en stock ({len(faltantes)}): {faltantes[:20]}{'...' if len(faltantes) > 20 else ''}")
    if not args.sin_formula_g:
        print(
            f"Fórmula G: costo neto sin II actualizado en {st_g['actualizados']} filas "
            f"(sin D: {st_g['sin_d']}, sin E: {st_g['sin_e']})"
        )
    if st_pv is not None:
        print(
            f"Precios venta: actualizados {st_pv['actualizados']} filas "
            f"(códigos CSV/Excel sin fila en DB: {st_pv['sin_producto']})"
        )


if __name__ == "__main__":
    main()
