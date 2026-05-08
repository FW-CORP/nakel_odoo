#!/usr/bin/env python3
"""
Exporta ``impuestos.sqlite`` a Excel (``.xlsx``): hoja ``productos`` y hoja ``meta``.

  python3 exportar_impuestos_sqlite_a_excel.py
  python3 exportar_impuestos_sqlite_a_excel.py --db /ruta/impuestos.sqlite --salida /ruta/salida.xlsx
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("Falta openpyxl. Ej.: pip install openpyxl") from exc

_DIR = Path(__file__).resolve().parent
_DEFAULT_DB = _DIR / "impuestos.sqlite"


def _autosize_columns(ws, max_width: int = 50) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max(len(str(c.value or "")) for c in col), max_width)
        ws.column_dimensions[letter].width = width + 2


def exportar(db_path: Path, salida: Path) -> None:
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row

    wb = openpyxl.Workbook()
    ws_p = wb.active
    ws_p.title = "productos"

    cur = con.execute("PRAGMA table_info(productos)")
    col_info = cur.fetchall()
    headers = [c[1] for c in col_info]
    ws_p.append(headers)
    for row in con.execute("SELECT * FROM productos ORDER BY referencia_interna"):
        ws_p.append(list(row))
    _autosize_columns(ws_p)

    ws_m = wb.create_sheet("meta")
    ws_m.append(["clave", "valor"])
    try:
        for row in con.execute("SELECT clave, valor FROM meta ORDER BY clave"):
            ws_m.append(list(row))
    except sqlite3.OperationalError:
        ws_m.append(["(sin tabla meta)", ""])
    _autosize_columns(ws_m)

    con.close()
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)


def main() -> int:
    ap = argparse.ArgumentParser(description="Exporta impuestos.sqlite a Excel")
    ap.add_argument("--db", type=Path, default=_DEFAULT_DB)
    ap.add_argument(
        "--salida",
        type=Path,
        default=None,
        help="Archivo .xlsx (default: mismo directorio que la DB, nombre con timestamp)",
    )
    args = ap.parse_args()

    if not args.db.is_file():
        print(f"No existe: {args.db}", file=sys.stderr)
        return 1

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    salida = args.salida or (args.db.parent / f"impuestos_export_{ts}.xlsx")

    if salida.suffix.lower() not in (".xlsx",):
        salida = salida.with_suffix(".xlsx")

    exportar(args.db, salida)
    print(f"Exportado: {salida.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
