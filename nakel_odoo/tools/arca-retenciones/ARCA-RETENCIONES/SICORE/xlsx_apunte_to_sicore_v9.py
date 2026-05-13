#!/usr/bin/env python3
"""
Convierte un Excel de **Apunte contable (account.move.line)** exportado desde Odoo
(columnas en español como en Inventario → Apuntes) a TXT listo para importación:

- **`sicore159`** (default): **SICORE v9 retenciones** — **159** posiciones, solo dígitos en montos,
  igual que `generar_sicore_v9_retenciones.py`.
- **`rgan145`**: **RGAN_CPA / SIAP** — **145** caracteres por línea, montos con **coma**
  decimal, mismo criterio que `SICORE/tools/generar_rgan_cpa_master_dev.py` (rama sin
  facturas reconciliadas: nro. orden desde el nombre de pago, importe total = retención
  si no pasás otra columna).

**¿Cuál usar?** Si el contador o SIAP piden el archivo tipo `RGAN_CPA*.TXT` del estudio
(con comas, largo 145), usá **`--formato rgan145`**. Si la grilla es la de importación
SICORE v9 extendida (159, sin separadores de miles), dejá el default **`sicore159`** o
pasalo explícito. La forma más segura es comparar con una **muestra aceptada** (largo
de línea: ¿145 o 159? ¿lleva comas en los montos?).

No usa Odoo ni XML-RPC: solo el archivo `.xlsx`.

Requisito: `openpyxl` (ej. `python3 -m venv .venv && .venv/bin/pip install openpyxl`).

Limitaciones del export típico de apuntes:
- No trae **CUIT** del partner: hay que pasar `--cuit-csv` (contacto → CUIT) o
  `--emitir-cuit-template` para generar la plantilla y completarla.
- No trae **base imponible** (`tax_base_amount`): por defecto va en **ceros**;
  usar `--base-igual-retencion` solo si el contador acepta ese criterio (no es lo
  mismo que la base real).
- No trae **importe del comprobante** (total OP/factura): por defecto se usa el
  mismo monto que la **retención** (`Crédito` o `Débito`), coherente con modo
  “solo OP / sin factura” en el generador desde Odoo.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    raise SystemExit(
        "Falta el paquete openpyxl. Creá un venv e instalá: pip install openpyxl\n"
        "Ejemplo: python3 -m venv /tmp/venv && /tmp/venv/bin/pip install openpyxl"
    ) from e

DEC2 = Decimal("0.01")
DEC3 = Decimal("0.001")
LREG = 159
LRGAN = 145


def validar_posiciones_clave_sicore(path: Path) -> tuple[int, list[tuple[int, list[str]]]]:
    """Igual criterio que `generar_sicore_v9_retenciones.py` (cols 67, 107, 109)."""
    raw = path.read_text(encoding="ascii", errors="replace")
    col_fecha, col_tipo, col_cuit = 67, 107, 109
    i_f, i_t, i_c = col_fecha - 1, col_tipo - 1, col_cuit - 1
    failures: list[tuple[int, list[str]]] = []
    n_ok = 0
    for lineno, line in enumerate(raw.splitlines(), 1):
        s = line.rstrip("\r\n")
        if not s:
            continue
        errs: list[str] = []
        if len(s) != LREG:
            errs.append(f"largo {len(s)} != {LREG}")
        if len(s) > i_f + 9:
            chunk = s[i_f : i_f + 10]
            if not (len(chunk) == 10 and chunk[2] == "/" and chunk[5] == "/"):
                errs.append(f"col {col_fecha}: esperaba dd/mm/aaaa, got {chunk!r}")
        else:
            errs.append(f"col {col_fecha}: línea corta")
        if len(s) > i_t + 1:
            if s[i_t : i_t + 2] != "80":
                errs.append(f"col {col_tipo}: esperaba '80', got {s[i_t:i_t+2]!r}")
        else:
            errs.append(f"col {col_tipo}: línea corta")
        if len(s) > i_c:
            if not s[i_c].isdigit():
                errs.append(f"col {col_cuit}: esperaba dígito, got {s[i_c]!r}")
        else:
            errs.append(f"col {col_cuit}: línea corta")
        if errs:
            failures.append((lineno, errs))
        else:
            n_ok += 1
    return n_ok, failures


def validar_rgan145_basico(path: Path) -> tuple[int, list[tuple[int, list[str]]]]:
    """Largo 145, prefijo 06, nro_orden (pos. 17–28) 12 dígitos, CUIT `80`+11 dígitos (pos. 110–122)."""
    raw = path.read_text(encoding="utf-8", errors="replace")
    failures: list[tuple[int, list[str]]] = []
    n_ok = 0
    for lineno, line in enumerate(raw.splitlines(), 1):
        s = line.rstrip("\r\n")
        if not s:
            continue
        errs: list[str] = []
        if len(s) != LRGAN:
            errs.append(f"largo {len(s)} != {LRGAN}")
        if not s.startswith("06"):
            errs.append(f"prefijo: esperaba '06', got {s[:2]!r}")
        nro = s[16:28] if len(s) >= 28 else ""
        if len(nro) != 12 or not nro.isdigit():
            errs.append(f"nro_orden (pos.17-28): esperaba 12 dígitos, got {nro!r}")
        cuit13 = s[109:122] if len(s) >= 122 else ""
        if len(cuit13) != 13:
            errs.append(f"cuit13: largo {len(cuit13)} != 13 ({cuit13!r})")
        elif not cuit13.startswith("80"):
            errs.append(f"cuit13: esperaba empezar con '80', got {cuit13!r}")
        elif not cuit13[2:].isdigit():
            errs.append(f"cuit13: esperaba 11 dígitos tras '80', got {cuit13!r}")
        if errs:
            failures.append((lineno, errs))
        else:
            n_ok += 1
    return n_ok, failures


def _d(x: Any) -> Decimal:
    if x is None or x is False:
        return Decimal("0")
    try:
        return Decimal(str(x))
    except Exception:
        return Decimal("0")


def _abs_money(x: Any) -> Decimal:
    return abs(_d(x)).quantize(DEC2, rounding=ROUND_HALF_UP)


def _digits_only(s: str) -> str:
    return re.sub(r"\D+", "", s or "")


def _ddmmyyyy(val: Any) -> str:
    if isinstance(val, datetime):
        return val.date().strftime("%d/%m/%Y")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    s = str(val or "").strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return datetime.strptime(s, "%Y-%m-%d").date().strftime("%d/%m/%Y")
    raise ValueError(f"Fecha no reconocida: {val!r}")


def _fmt_centavos_sin_separador(value: Decimal, width: int) -> str:
    cents = (value * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    try:
        n = int(cents)
    except Exception:
        n = 0
    if n < 0:
        n = -n
    s = str(n).zfill(width)
    return s[-width:]


def _nro_comprobante_16_fiscal(raw: str) -> str:
    s = (raw or "").strip()
    if not s:
        return "0" * 16
    m = re.match(r"^\s*(\d{1,9})\s*[-/]\s*(\d{1,12})\s*$", s)
    if m:
        pv = int(m.group(1))
        num = int(m.group(2))
        pv = max(0, min(pv, 9999))
        if num <= 99999999:
            return f"0000{pv:04d}{num:08d}"
        num = max(0, min(num, 999999999999))
        return f"{pv:04d}{num:012d}"
    dig = _digits_only(s)
    if not dig:
        return "0" * 16
    if len(dig) >= 16:
        return dig[-16:]
    if len(dig) >= 10:
        pv = int(dig[:4])
        rest = dig[4:]
        num = int(rest) if rest else 0
        return f"{pv:04d}{num:012d}"
    return dig.zfill(16)[-16:]


def _cod_regimen_3_desde_texto_impuesto(impuesto: str) -> str:
    """Ej.: 'Retención gcias 830 sobre Bienes' → '830'."""
    s = str(impuesto or "")
    m = re.search(r"(?i)gcias\s+(\d{3})\b", s)
    if m:
        return m.group(1).zfill(3)
    m2 = re.search(r"\b(\d{3})\s+sobre\b", s, flags=re.I)
    if m2:
        return m2.group(1).zfill(3)
    d = _digits_only(s)
    if len(d) >= 3:
        return d[-3:].zfill(3)
    return "000"


def _fmt_cuit_11(cuit11: str) -> str:
    d = _digits_only(cuit11)
    if len(d) > 11:
        d = d[-11:]
    elif len(d) < 11:
        d = d.zfill(11)
    return d[:11]


def _cuit_desde_contacto(contacto: str) -> str | None:
    s = str(contacto or "")
    for m in re.finditer(r"\d{11}", _digits_only(s)):
        return m.group(0)
    return None


def _reg_sicore_line_160(
    *,
    cod_comprobante_2: str,
    fecha_comprobante_10: str,
    nro_comprobante_16: str,
    importe_comprobante_16: str,
    cod_impuesto_4: str,
    cod_regimen_3: str,
    cod_operacion_1: str,
    base_14: str,
    fecha_retencion_10: str,
    cod_condicion_2: str,
    importe_retencion_14: str,
    excedente_otros_14: str,
    tipo_doc_2: str,
    nro_cuit_11: str,
    porc_exclusion_11: str,
    fecha_pub_certificado_10: str,
    tipo_regimen_especial_2: str,
    importe_base_exclusion_14: str,
    tipo_cuenta_2: str,
    relleno_final_1: str,
) -> str:
    core = (
        cod_comprobante_2
        + fecha_comprobante_10
        + nro_comprobante_16
        + importe_comprobante_16
        + cod_impuesto_4
        + cod_regimen_3
        + cod_operacion_1
        + base_14
        + fecha_retencion_10
        + cod_condicion_2
        + importe_retencion_14
        + excedente_otros_14
        + tipo_doc_2
        + nro_cuit_11
        + porc_exclusion_11
    )
    tail = (
        fecha_pub_certificado_10
        + tipo_regimen_especial_2
        + importe_base_exclusion_14
        + tipo_cuenta_2
        + relleno_final_1
    )
    out = core + tail
    if len(out) != LREG:
        raise ValueError(f"Largo registro {len(out)} != {LREG}")
    return out


def _header_map(ws: Any) -> dict[str, int]:
    headers: dict[str, int] = {}
    for c in range(1, 300):
        v = ws.cell(1, c).value
        if v is None:
            if c > 25 and not any(ws.cell(1, cc).value for cc in range(c, c + 5)):
                break
            continue
        headers[str(v).strip()] = c
    return headers


def _col(headers: dict[str, int], name: str, aliases: tuple[str, ...] = ()) -> int:
    for k in (name,) + aliases:
        if k in headers:
            return headers[k]
    raise SystemExit(
        f"No se encontró la columna {name!r} en la fila 1 del Excel. "
        f"Encabezados: {sorted(headers)!r}"
    )


def _load_cuit_map(path: Path) -> dict[str, str]:
    """
    CSV UTF-8: columnas `contacto` y `cuit` (cabecera opcional).
    También acepta dos columnas sin cabecera: contacto;cuit
    """
    raw = path.read_text(encoding="utf-8-sig")
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(raw[:4096])
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(raw.splitlines(), dialect=dialect))
    if not rows:
        return {}
    out: dict[str, str] = {}
    start = 0
    if rows[0] and rows[0][0].lower().strip() in ("contacto", "partner", "nombre"):
        start = 1
    for parts in rows[start:]:
        if len(parts) < 2:
            continue
        name = parts[0].strip()
        cuit = _fmt_cuit_11(parts[1])
        if name:
            out[name.casefold()] = cuit
    return out


def _emit_cuit_template(xlsx: Path, out_csv: Path) -> None:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = _header_map(ws)
    ic = _col(headers, "Contacto")
    partners: list[str] = []
    seen: set[str] = set()
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, ic).value or "").strip()
        if name and name.casefold() not in seen:
            seen.add(name.casefold())
            partners.append(name)
    wb.close()
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["contacto", "cuit"])
        for p in sorted(partners, key=str.casefold):
            w.writerow([p, ""])
    print(f"Plantilla CUIT escrita: {out_csv} ({len(partners)} filas). Completá la columna cuit.")


def _retencion_monto(row: dict[str, Any]) -> Decimal:
    cred = _abs_money(row.get("credito"))
    deb = _abs_money(row.get("debito"))
    if cred > 0:
        return cred
    if deb > 0:
        return deb
    return _abs_money(row.get("balance"))


def _abs_money_dec3(x: Any) -> Decimal:
    return abs(_d(x)).quantize(DEC3, rounding=ROUND_HALF_UP)


def _nro_orden_12_rgan_solo_digitos(raw: str, fallback_id: int) -> str:
    """Pos. 17–28 RGAN: 12 dígitos (mismo criterio que `generar_rgan_cpa_master_dev.py`)."""
    d = _digits_only(str(raw or ""))
    if d:
        return d.zfill(12)[-12:]
    try:
        fid = int(fallback_id)
    except Exception:
        fid = 0
    return str(fid).zfill(12)[-12:]


def _fmt_num_coma(x: Decimal, *, decimals: int) -> str:
    q = DEC3 if decimals == 3 else DEC2
    d = x.quantize(q, rounding=ROUND_HALF_UP)
    s = f"{d:.{decimals}f}"
    return s.replace(".", ",")


def _fmt_field_right(value: str, width: int) -> str:
    return (value or "").rjust(width)[:width]


def _fmt_codigo8_like_marzo(code: str) -> str:
    """Replica `generar_rgan_cpa_master_dev.py` para el valor típico del estudio."""
    s = str(code or "").strip()
    if s.isdigit() and int(s) == 2170781:
        return "2170781 "
    return s.zfill(8)[:8]


def _fmt_date_rgan(val: Any) -> str:
    return _ddmmyyyy(val).ljust(10)[:10]


def _build_line_rgan145(
    *,
    fecha_val: Any,
    sucursal: str,
    raw_numero: str,
    fallback_nro: int,
    imp_total: Decimal,
    base: Decimal,
    reten: Decimal,
    cuit11: str,
    codigo_8: str,
    jurisd_3: str,
) -> str:
    fecha = _fmt_date_rgan(fecha_val)
    fecha2 = fecha
    fecha3 = fecha
    nro_orden = _nro_orden_12_rgan_solo_digitos(raw_numero, fallback_nro)
    imp_total_s = _fmt_field_right(_fmt_num_coma(_abs_money_dec3(imp_total), decimals=3), 12)
    base_s = _fmt_field_right(_fmt_num_coma(base, decimals=2), 13)
    reten_s = _fmt_field_right(_fmt_num_coma(reten, decimals=2), 14)
    c11 = _digits_only(cuit11)
    if len(c11) > 11:
        c11 = c11[-11:]
    cuit13 = ("80" + c11) if len(c11) == 11 else "80".ljust(13, "0")
    line = (
        "06"
        + fecha
        + str(sucursal).zfill(4)[:4]
        + nro_orden
        + (" " * 5)
        + imp_total_s
        + _fmt_codigo8_like_marzo(codigo_8)
        + base_s
        + fecha2
        + str(jurisd_3).zfill(3)[:3]
        + reten_s
        + (" " * 2)
        + "0,00"
        + fecha3
        + cuit13[:13]
        + (" " * 9)
        + ("0" * 14)
    )
    return line[:LRGAN].ljust(LRGAN)


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", type=Path, required=True, help="Ruta al Excel de apuntes")
    ap.add_argument("--sheet", default=None, help="Nombre de hoja (default: la primera)")
    ap.add_argument(
        "--formato",
        choices=("rgan145", "sicore159"),
        default="sicore159",
        help="rgan145=RGAN_CPA/SIAP (145 chars, comas); sicore159=SICORE v9 importación (159, solo dígitos). Default sicore159.",
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=None,
        help="TXT de salida. Obligatorio salvo --emitir-cuit-template.",
    )
    ap.add_argument(
        "--emitir-cuit-template",
        type=Path,
        metavar="CSV",
        help="Solo genera CSV contacto;cuit vacío y termina (no escribe TXT).",
    )
    ap.add_argument("--cuit-csv", type=Path, help="CSV contacto → CUIT (ver docstring)")
    ap.add_argument(
        "--omitir-sin-cuit",
        action="store_true",
        help="No falla: omite filas sin CUIT (lista advertencias en stderr).",
    )
    ap.add_argument("--codigo-comprobante", default="06", help="Tabla A, 2 dígitos (default 06 OP)")
    ap.add_argument("--codigo-impuesto", default="0217", help="4 dígitos Ganancias")
    ap.add_argument(
        "--codigo-operacion",
        default="1",
        help="Tabla C ARCA: 1=Retención (criterio estudio; el script Odoo a veces usa 0).",
    )
    ap.add_argument("--codigo-condicion", default="01", help="Tabla D, 2 dígitos")
    ap.add_argument("--tipo-doc", default="80", help="Tabla F: 80=CUIT")
    ap.add_argument(
        "--base-igual-retencion",
        action="store_true",
        help="Rellena base imponible = importe retención (solo si el contador lo autoriza).",
    )
    ap.add_argument(
        "--validar",
        action="store_true",
        help="Tras escribir: si formato rgan145 comprueba largo 145 y CUIT; si sicore159, cols 67/107/109.",
    )
    ap.add_argument(
        "--importe-comprobante-columna",
        default=None,
        help="Columna numérica del Excel para importe comprobante / total OP (sicore159); en rgan145 es el importe total del bloque de 12 (3 decimales).",
    )
    ap.add_argument(
        "--sucursal",
        default="0001",
        help="Solo rgan145: código sucursal 4 caracteres (default 0001).",
    )
    ap.add_argument(
        "--codigo-8-rgan",
        default="02170781",
        dest="codigo_8_rgan",
        help="Solo rgan145: campo de 8 posiciones (default 02170781; 2170781 usa formato del estudio).",
    )
    ap.add_argument(
        "--jurisd",
        default="010",
        help="Solo rgan145: jurisdicción 3 caracteres (default 010).",
    )
    args = ap.parse_args(argv)

    xlsx = args.xlsx.expanduser().resolve()
    if not xlsx.is_file():
        raise SystemExit(f"No existe el Excel: {xlsx}")

    if args.emitir_cuit_template:
        _emit_cuit_template(xlsx, args.emitir_cuit_template.expanduser().resolve())
        return 0

    if not args.out:
        raise SystemExit("Falta --out (ruta del TXT de salida).")

    cuit_map: dict[str, str] = {}
    if args.cuit_csv:
        p = args.cuit_csv.expanduser().resolve()
        if not p.is_file():
            raise SystemExit(f"No existe --cuit-csv: {p}")
        cuit_map = _load_cuit_map(p)

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[args.sheet or wb.sheetnames[0]]
    headers = _header_map(ws)
    cols = {
        "contacto": _col(headers, "Contacto"),
        "cuenta": _col(headers, "Cuenta"),
        "credito": _col(headers, "Crédito", ("Credito",)),
        "debito": _col(headers, "Débito", ("Debito",)),
        "balance": _col(headers, "Balance"),
        "fecha": _col(headers, "Fecha"),
        "numero": _col(headers, "Número", ("Numero",)),
        "impuesto": _col(headers, "Impuesto del emisor"),
    }
    col_imp_comp: int | None = None
    if args.importe_comprobante_columna:
        col_imp_comp = _col(headers, args.importe_comprobante_columna)

    fecha_pub_def = "00/00/0000"
    tipo_reg_esp_def = "00"
    base_excl_def = "0" * 14
    tipo_cta_def = "00"
    relleno_final_def = " "

    out_lines: list[str] = []
    skipped: list[str] = []

    for r in range(2, ws.max_row + 1):
        contacto = str(ws.cell(r, cols["contacto"]).value or "").strip()
        cuenta = str(ws.cell(r, cols["cuenta"]).value or "")
        impuesto = str(ws.cell(r, cols["impuesto"]).value or "")
        if "SICORE" not in cuenta and "retención" not in impuesto.lower() and "retencion" not in impuesto.lower():
            continue

        ret = _retencion_monto(
            {
                "credito": ws.cell(r, cols["credito"]).value,
                "debito": ws.cell(r, cols["debito"]).value,
                "balance": ws.cell(r, cols["balance"]).value,
            }
        )
        if ret == 0:
            continue

        fecha_cell = ws.cell(r, cols["fecha"]).value
        fecha_ret = _ddmmyyyy(fecha_cell)
        fecha_comp = fecha_ret
        raw_nro = str(ws.cell(r, cols["numero"]).value or "").strip()
        nro_16 = _nro_comprobante_16_fiscal(raw_nro)

        cuit_raw = cuit_map.get(contacto.casefold(), "") or _cuit_desde_contacto(contacto) or ""
        if not cuit_raw:
            msg = f"Fila {r}: sin CUIT para contacto {contacto!r}"
            if args.omitir_sin_cuit:
                skipped.append(msg)
                continue
            raise SystemExit(msg + "\nUsá --cuit-csv, --emitir-cuit-template o --omitir-sin-cuit.")

        if col_imp_comp is not None:
            imp_comp = _abs_money(ws.cell(r, col_imp_comp).value)
        else:
            imp_comp = ret

        base_dec = ret if args.base_igual_retencion else Decimal("0")

        if args.formato == "sicore159":
            base_s = _fmt_centavos_sin_separador(base_dec, 14)
            imp_ret_s = _fmt_centavos_sin_separador(ret, 14)
            imp_comp_s = _fmt_centavos_sin_separador(imp_comp, 16)
            cod_reg = _cod_regimen_3_desde_texto_impuesto(impuesto)
            line = _reg_sicore_line_160(
                cod_comprobante_2=str(args.codigo_comprobante).zfill(2)[:2],
                fecha_comprobante_10=fecha_comp,
                nro_comprobante_16=nro_16,
                importe_comprobante_16=imp_comp_s,
                cod_impuesto_4=str(args.codigo_impuesto).zfill(4)[:4],
                cod_regimen_3=cod_reg,
                cod_operacion_1=str(args.codigo_operacion)[:1],
                base_14=base_s,
                fecha_retencion_10=fecha_ret,
                cod_condicion_2=str(args.codigo_condicion).zfill(2)[:2],
                importe_retencion_14=imp_ret_s,
                excedente_otros_14="0" * 14,
                tipo_doc_2=str(args.tipo_doc).zfill(2)[:2],
                nro_cuit_11=_fmt_cuit_11(cuit_raw),
                porc_exclusion_11="0" * 11,
                fecha_pub_certificado_10=fecha_pub_def,
                tipo_regimen_especial_2=tipo_reg_esp_def,
                importe_base_exclusion_14=base_excl_def,
                tipo_cuenta_2=tipo_cta_def,
                relleno_final_1=relleno_final_def,
            )
        else:
            line = _build_line_rgan145(
                fecha_val=fecha_cell,
                sucursal=args.sucursal,
                raw_numero=raw_nro,
                fallback_nro=r,
                imp_total=imp_comp,
                base=base_dec,
                reten=ret,
                cuit11=_fmt_cuit_11(cuit_raw),
                codigo_8=args.codigo_8_rgan,
                jurisd_3=args.jurisd,
            )
        out_lines.append(line)

    wb.close()

    if skipped:
        for s in skipped:
            print(s, file=sys.stderr)

    if not out_lines:
        raise SystemExit("No se generó ninguna línea (¿filtro de cuenta/impuesto o todo sin CUIT?).")

    out_path = args.out.expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    enc = "utf-8" if args.formato == "rgan145" else "ascii"
    with out_path.open("w", encoding=enc, errors="replace", newline="") as f:
        for ln in out_lines:
            f.write(ln + "\r\n")

    print(f"OK: {out_path} ({len(out_lines)} líneas, formato={args.formato}).")
    if args.validar:
        if args.formato == "sicore159":
            n_ok, failures = validar_posiciones_clave_sicore(out_path)
            label = "SICORE 159"
        else:
            n_ok, failures = validar_rgan145_basico(out_path)
            label = "RGAN/SIAP 145"
        if failures:
            print(f"VALIDACIÓN FALLO ({label}): {len(failures)} línea(s), {n_ok} OK")
            for lineno, errs in failures[:10]:
                print(f"  Línea {lineno}: {errs}")
            return 1
        print(f"VALIDACIÓN OK ({label}): {n_ok} línea(s).")
    else:
        print("Validación opcional: re-ejecutá con --validar")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
