#!/usr/bin/env python3
"""
Arma `RGAN_CPA` / SIAP (145 + CRLF) desde:
- Excel de apuntes (mismas columnas que `xlsx_apunte_to_sicore_v9.py`)
- TXT modelo ya validado en SIAP: se reutiliza la **línea completa** cuando coincide
  **fecha de retención** + **importe retenido** (clave única en la práctica de este set)
- Mapa contacto → CUIT (CSV) para filas sin match en el modelo (p. ej. salidas posteriores al rango del TXT)

Uso:
  python3 rgan145_desde_apunte_y_modelo.py \\
    --xlsx apunte.xlsx \\
    --modelo RGAN_CPA_SIAP_....TXT \\
    --cuit-csv cuits.csv \\
    --out salida.TXT
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    raise SystemExit("Instalá openpyxl (venv + pip install openpyxl).") from e

DEC2 = Decimal("0.01")
DEC3 = Decimal("0.001")
LRGAN = 145


def _d(x: Any) -> Decimal:
    if x is None or x is False:
        return Decimal("0")
    try:
        return Decimal(str(x))
    except Exception:
        return Decimal("0")


def _abs_money(x: Any) -> Decimal:
    return abs(_d(x)).quantize(DEC2, rounding=ROUND_HALF_UP)


def _abs_money_dec3(x: Any) -> Decimal:
    return abs(_d(x)).quantize(DEC3, rounding=ROUND_HALF_UP)


def _digits_only(s: str) -> str:
    return re.sub(r"\D+", "", s or "")


def _ddmmyyyy(val: Any) -> str:
    if isinstance(val, datetime):
        return val.date().strftime("%d/%m/%Y")
    s = str(val or "").strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return datetime.strptime(s, "%Y-%m-%d").date().strftime("%d/%m/%Y")
    raise ValueError(f"Fecha: {val!r}")


def _header_map(ws: Any) -> dict[str, int]:
    headers: dict[str, int] = {}
    for c in range(1, 300):
        v = ws.cell(1, c).value
        if v is None:
            if c > 25 and not any(ws.cell(1, cc).value for cc in range(c, min(c + 5, 300))):
                break
            continue
        headers[str(v).strip()] = c
    return headers


def _col(headers: dict[str, int], name: str, aliases: tuple[str, ...] = ()) -> int:
    for k in (name,) + aliases:
        if k in headers:
            return headers[k]
    raise SystemExit(f"No columna {name!r}. Hay: {sorted(headers)!r}")


def _load_cuit_map(path: Path) -> dict[str, str]:
    raw = path.read_text(encoding="utf-8-sig")
    try:
        dialect = csv.Sniffer().sniff(raw[:4096])
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(raw.splitlines(), dialect=dialect))
    out: dict[str, str] = {}
    start = 0
    if rows and rows[0] and rows[0][0].lower().strip() in ("contacto", "partner", "nombre"):
        start = 1
    for parts in rows[start:]:
        if len(parts) < 2:
            continue
        name, cuit = parts[0].strip(), _digits_only(parts[1])
        if len(cuit) > 11:
            cuit = cuit[-11:]
        elif len(cuit) < 11:
            cuit = cuit.zfill(11)
        if name:
            out[name.casefold()] = cuit[:11]
    return out


def _parse_model_line(line: str) -> dict[str, Any] | None:
    s = line.rstrip("\r\n")
    if len(s) != LRGAN:
        return None
    fecha2 = s[66:76].strip()
    ret_s = s[79:93].strip().replace(".", "").replace(",", ".")
    try:
        ret = Decimal(ret_s).quantize(DEC2, rounding=ROUND_HALF_UP)
    except Exception:
        return None
    return {"line": s, "fecha2": fecha2, "ret": ret}


def _nro_orden_12_rgan_solo_digitos(raw: str, fallback_id: int) -> str:
    """Pos. 17–28 RGAN: 12 dígitos (mismo criterio que `generar_rgan_cpa_master_dev.py`)."""
    d = _digits_only(str(raw or ""))
    if d:
        return d.zfill(12)[-12:]
    try:
        fid = int(fallback_id)
    except Exception:
        fid = 0
    return str(fid).zfill(12)[-12:]


def _fmt_num_coma(x: Decimal, *, decimals: int) -> str:
    q = DEC3 if decimals == 3 else DEC2
    d = x.quantize(q, rounding=ROUND_HALF_UP)
    s = f"{d:.{decimals}f}"
    return s.replace(".", ",")


def _fmt_field_right(value: str, width: int) -> str:
    return (value or "").rjust(width)[:width]


def _fmt_codigo8_like_marzo(code: str) -> str:
    s = str(code or "").strip()
    if s.isdigit() and int(s) == 2170781:
        return "2170781 "
    return s.zfill(8)[:8]


def _build_line_rgan145(
    *,
    fecha_val: Any,
    sucursal: str,
    raw_numero: str,
    fallback_nro: int,
    imp_total: Decimal,
    base: Decimal,
    reten: Decimal,
    cuit11: str,
    codigo_8: str,
    jurisd_3: str,
) -> str:
    fecha = _ddmmyyyy(fecha_val).ljust(10)[:10]
    fecha2 = fecha
    fecha3 = fecha
    nro_orden = _nro_orden_12_rgan_solo_digitos(raw_numero, fallback_nro)
    imp_total_s = _fmt_field_right(_fmt_num_coma(_abs_money_dec3(imp_total), decimals=3), 12)
    base_s = _fmt_field_right(_fmt_num_coma(base, decimals=2), 13)
    reten_s = _fmt_field_right(_fmt_num_coma(reten, decimals=2), 14)
    c11 = _digits_only(cuit11)
    if len(c11) > 11:
        c11 = c11[-11:]
    cuit13 = ("80" + c11) if len(c11) == 11 else "80".ljust(13, "0")
    line = (
        "06"
        + fecha
        + str(sucursal).zfill(4)[:4]
        + nro_orden
        + (" " * 5)
        + imp_total_s
        + _fmt_codigo8_like_marzo(codigo_8)
        + base_s
        + fecha2
        + str(jurisd_3).zfill(3)[:3]
        + reten_s
        + (" " * 2)
        + "0,00"
        + fecha3
        + cuit13[:13]
        + (" " * 9)
        + ("0" * 14)
    )
    return line[:LRGAN].ljust(LRGAN)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", type=Path, required=True)
    ap.add_argument("--modelo", type=Path, required=True, help="TXT RGAN 145 del estudio/SIAP OK")
    ap.add_argument("--cuit-csv", type=Path, required=True)
    ap.add_argument("--out", type=Path, required=True)
    ap.add_argument("--sucursal", default="0001")
    ap.add_argument("--codigo-8", default="02170781", dest="codigo_8")
    ap.add_argument("--jurisd", default="010")
    args = ap.parse_args()

    modelo_path = args.modelo.expanduser().resolve()
    raw_model = modelo_path.read_text(encoding="utf-8", errors="replace")
    pool: dict[tuple[str, Decimal], list[str]] = defaultdict(list)
    for ln in raw_model.splitlines():
        p = _parse_model_line(ln)
        if not p:
            continue
        pool[(p["fecha2"], p["ret"])].append(p["line"])

    cuit_map = _load_cuit_map(args.cuit_csv.expanduser().resolve())

    xlsx = args.xlsx.expanduser().resolve()
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = _header_map(ws)
    cols = {
        "contacto": _col(headers, "Contacto"),
        "cuenta": _col(headers, "Cuenta"),
        "credito": _col(headers, "Crédito", ("Credito",)),
        "debito": _col(headers, "Débito", ("Debito",)),
        "balance": _col(headers, "Balance"),
        "fecha": _col(headers, "Fecha"),
        "numero": _col(headers, "Número", ("Numero",)),
        "impuesto": _col(headers, "Impuesto del emisor"),
    }

    out_lines: list[str] = []
    desde_modelo = 0
    sintetico = 0

    for r in range(2, ws.max_row + 1):
        cuenta = str(ws.cell(r, cols["cuenta"]).value or "")
        impuesto = str(ws.cell(r, cols["impuesto"]).value or "")
        if "SICORE" not in cuenta and "retención" not in impuesto.lower() and "retencion" not in impuesto.lower():
            continue

        cred = ws.cell(r, cols["credito"]).value
        deb = ws.cell(r, cols["debito"]).value
        bal = ws.cell(r, cols["balance"]).value
        ret = _abs_money(cred) if _abs_money(cred) > 0 else (_abs_money(deb) if _abs_money(deb) > 0 else _abs_money(bal))
        if ret == 0:
            continue

        fecha_cell = ws.cell(r, cols["fecha"]).value
        fe = _ddmmyyyy(fecha_cell)
        contacto = str(ws.cell(r, cols["contacto"]).value or "").strip()
        raw_nro = str(ws.cell(r, cols["numero"]).value or "").strip()

        key = (fe, ret)
        if pool.get(key):
            out_lines.append(pool[key].pop(0))
            desde_modelo += 1
            continue

        c11 = cuit_map.get(contacto.casefold())
        if not c11:
            raise SystemExit(f"Fila {r}: sin línea en modelo para ({fe!r}, {ret}) y sin CUIT en CSV para {contacto!r}")

        line = _build_line_rgan145(
            fecha_val=fecha_cell,
            sucursal=args.sucursal,
            raw_numero=raw_nro,
            fallback_nro=r,
            imp_total=ret,
            base=Decimal("0"),
            reten=ret,
            cuit11=c11,
            codigo_8=args.codigo_8,
            jurisd_3=args.jurisd,
        )
        out_lines.append(line)
        sintetico += 1

    wb.close()

    out_path = args.out.expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        for ln in out_lines:
            f.write(ln + "\r\n")

    print(
        f"OK: {out_path} ({len(out_lines)} líneas). "
        f"Copiadas del modelo: {desde_modelo}; sintéticas (Excel+CSV): {sintetico}."
    )
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
